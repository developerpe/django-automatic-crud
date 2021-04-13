from datetime import datetime

from django.http import HttpResponse
from django.views.generic import TemplateView

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,Border,Font,Side
)
try:
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter

from automatic_crud.utils import (
    get_model,get_model_fields_names,get_queryset
)

def excel_report_title(__model_name: str):
    date = datetime.now()
    title = "REPORTE DE {0} EN FORMATO EXCEL REALIZADO EN LA FECHA: {1}".format(
                                                                            __model_name.upper(),
                                                                            "%s/%s/%s" % (
                                                                                date.day,date.month,
                                                                                date.year
                                                                            )
                                                                        )
    return title

def validate_id(__field: str):
    if str(__field).lower() != 'id':
        return True
    return False

class ExcelReportFormat:
    """ FORMATO PARA REPORTE EN EXCEL DE CUALQUIER MODELO.

    Esta clase realiza un formato base para generar un reporte en excel para cualquier
    modelo que se desee, en el constructor se definen los parámetros a utilizar, además
    se cuenta con algunos métodos que construyen bloque a bloque el reporte.

    Funciones:
    __init__                        -- función constructora de inicio, en ella se recibe el nombre de la aplicacion
                                       donde se encuentra el modelo a utilizar y el nombre del modelo a utilizar.

        Parámetros:
        _app_name                   -- nombre de aplicacion donde se encuentra el modelo a utilizar.
        _model_name                 -- nombre de tipo Texto de modelo a utilizar.

        Variables:
        _app_name                   -- nombre de aplicacion donde se encuentra el modelo a utilizar.
        _model_name                 -- nombre de modelo a utilizar.
        _modelo                     -- modelo en cuestión a utilizarse, se obtiene por medio de una función
                                       que busca en todo el proyecto el modelo con los parámetros: _app_name
                                       y _model_name.
        _nombres_atributos_modelos  -- lista  con los nombres de todos los atributos que
                                       contiene el modelo a excepción de la clave primaria.
        _consulta                   -- consulta que contiene todos los registros del modelo en cuestión.
        _titulo_cabecera            -- titulo de reporte en excel creado a través de una funcipón que incluye
                                       el nombre del modelo y la fecha de creacion del reporte.
        _libro_trabajo              -- libro de trabajo actual en el que se creará el reporte, es una instancia
                                       de Workbook.
        _hoja_trabajo               -- hoja de trabajo inicial donde se pintará el reporte.

    cabecera_tabla_reporte_excel    -- función que recibe opcionalmente como parámetro la dimensión o altura de las
                                       filas y columnas, y genera todo el encabezado de la tabla en excel.
                                       Por defecto se empezara a pintar la cabecera de la tabla en la fila número 3
                                       y el título desde la celda B1.
                                       La cabecera de la tabla empieza en la letra A y su control de calcula automaticamente
                                       por el parámetro cont, se pinta cada encabezado tomando los valores guardados en la
                                       lista _nombres_atributos_modelos.

        Parámetros:
        dimension_fila              -- parámetro opcional definido por defecto en 15 que indica la altura de las filas.
        dimension_columna           -- parámetro opcional definido por defecto en 25 que indica la altura de las columnas.

        Variables:
        letra_cabecera              -- letra formada automaticamente y que indica el final de la celda que se uniran o merguearán
                                       para ubicar correctamente el titulo de la misma.

    pintar_valores_excel            -- función que pinta los valores correspondientes a la consulta del modelo en cuestion, guardada en la
                                       variable _consulta, se recorre el diccionario y se procede a validar que el campo no corresponda al
                                       atributo id ya que esta será la clave primaria y no se desea renderizar.
                                       Se hace una validación cuando el tipo de dato es un Boolean para escribir True o False según corresponda.

    """
    

    def __init__(self,__app_name:str,__model_name:str, *args, **kwargs):
        self.__app_name = __app_name
        self.__model_name = __model_name
        self.__model = get_model(self.__app_name,self.__model_name)
        self.__model_fields_names = get_model_fields_names(self.__model)
        self.__queryset = get_queryset(self.__model)
        self.__report_title = excel_report_title(self.__model_name)
        self.__workbook = Workbook()
        self.__sheetwork = self.__workbook.active

    def __excel_report_header(self,row_dimension = 15, col_dimension = 25):
        self.__sheetwork['B1'].alignment = Alignment(horizontal = "center", vertical = "center")
        self.__sheetwork['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        self.__sheetwork['B1'].font = Font(name = 'Calibri', size = 12, bold = True)
        self.__sheetwork['B1'] = self.__report_title
        print(len(self.__model_fields_names))
        if len(self.__model_fields_names) < 12:
            __header_letter = 'L'
        else:
            __header_letter = '{0}'.format(get_column_letter(len(self.__model_fields_names)).upper())
        self.__sheetwork.merge_cells('B1:{0}1'.format(__header_letter))
        self.__sheetwork.row_dimensions[3].height = row_dimension

        __count = 1
        for __field in self.__model_fields_names:
            __letter = get_column_letter(__count).upper()
            self.__sheetwork['{0}3'.format(__letter)].alignment = Alignment(horizontal = "center", vertical = "center")
            self.__sheetwork['{0}3'.format(__letter)].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            self.__sheetwork['{0}3'.format(__letter)].font = Font(name = 'Calibri', size = 9, bold = True)
            self.__sheetwork['{0}3'.format(__letter)] = '{0}'.format(__field.upper())
            self.__sheetwork.column_dimensions['{0}'.format(__letter)].height = col_dimension
            __count += 1
    
    def __print_values(self):
        row_count = 4
        col_count = 1
        general_count = len(self.__model_fields_names)
        data = [value for value in self.__queryset]

        for value in data:
            for key,subvalue in value.items():
                if validate_id(key):
                    if key not in self.__model.exclude_fields:
                        if general_count > 0:
                            self.__sheetwork.cell(row = row_count, column = col_count).alignment = Alignment(horizontal = "center")
                            self.__sheetwork.cell(row = row_count, column = col_count).border = Border(left = Side(border_style = "thin"),
                                                                        right = Side(border_style = "thin"),top = Side(border_style = "thin"), 
                                                                        bottom = Side(border_style = "thin"))
                            if type(subvalue) is bool:
                                if subvalue is True:
                                    self.__sheetwork.cell(row = row_count, column = col_count).value = 'No eliminado'
                                else:
                                    self.__sheetwork.cell(row = row_count, column = col_count).value = 'Eliminado'
                            else:
                                self.__sheetwork.cell(row = row_count, column = col_count).value = str(subvalue)
                                if (self.__sheetwork.column_dimensions[get_column_letter(col_count).upper()].width < len(str(subvalue))):
                                    self.__sheetwork.column_dimensions[get_column_letter(col_count).upper()].width = len(str(subvalue))
                            col_count += 1
                            general_count -= 1

                    if general_count == 0:
                        general_count = len(self.__model_fields_names)
            
            row_count += 1
            col_count = 1

    def get_excel_report(self):
        report_name = "Reporte {0} en Excel .xlsx".format(self.__model_name)
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(report_name)
        response['Content-Disposition'] = content
        self.__workbook.save(response)
        return response

    def build_report(self):
        self.__excel_report_header()
        self.__print_values()

class GetExcelReport(TemplateView):
    def get(self,request,_app_name:str,_model_name:str,*args,**kwargs):
        __report = ExcelReportFormat(_app_name,_model_name)
        __report.build_report()
        return __report.get_excel_report()